from openpyxl import load_workbook
from pymongo import MongoClient
from datetime import date, datetime, timedelta

#Record Start time
start_time = datetime.now()

CONNECTION_STRING= "mongodb+srv://prod1:mismis2020@cluster0.r1zhb.mongodb.net"
client = MongoClient (CONNECTION_STRING)

template = "F:/mmgapp/dev/booking/housekeeping/template/QC DispatchReport.xlsm"

output = "F:/mmgapp/dev/booking/housekeeping/dispatchReport/"

db = client["dev_qcDB"]
inspectionBooking = db["inspectionBooking"]
metaTable = db["metaTable"]
factoryMaster = db["factoryMaster"]
partyTable = db["partyTable"]
mcTable = db["mcTable"]

## Get All QA Lead in Factory Master
factoryMasterData = list(factoryMaster.find({}))

allQaLeads = list({item['qa_lead'] for item in factoryMasterData if item['qa_lead'] != ''})

## Calculate report period
today = date.today()
reportStart = (today + timedelta(days=(7 - today.weekday() - 1)))
reportEnd = ((today + timedelta(days=(6 - today.weekday() + 6))) + timedelta(days=7))
## For Output
reportStartStr = (today + timedelta(days=(7 - today.weekday() - 1))).strftime("%m%d%Y")
reportEndStr = ((today + timedelta(days=(6 - today.weekday() + 6))) + timedelta(days=7)).strftime("%m%d%Y")

## Get Booking Record in report period
bookingRecord = list(inspectionBooking.find({"main.inspection_date": {"$gte": reportStart.strftime("%Y-%m-%d"),"$lte": reportEnd.strftime("%Y-%m-%d") } }))

## Covert connection to list data to avoid repeat loading in loop
factoryMasterData = factoryMasterData
inspecTypeData = list(metaTable.find_one({'category': "inspType"})["selectionList"])
countryData = list(metaTable.find_one({'category': "countryRegion"})["selectionList"])
partyTableData = list(partyTable.find({}))
mcTableData = list(mcTable.find({}))

for qaLead in allQaLeads:

    # Active workbook
    wb = load_workbook(template, keep_vba=True)

    # File name
    row_index = 2
    record_index = 3
    outputFile = output + "QC DispatchReport - " + str(qaLead) + "(" + str(reportStartStr) + "-" + str(reportEndStr) + ").xlsm"

    for rec in bookingRecord:
        
        # initize
        orderQty = 0
        shipQty = 0

        # Search booking record QA Lead by using SU, MF number
        qaLeadInRecord = next((record['qa_lead'] for record in factoryMasterData if (record['su_no'] == rec['main']['su_no']) and (record['mf_no'] == rec['main']['mf_no'])), "")
            
        if qaLeadInRecord == qaLead:

            row_index = row_index + 1
                
            for item in rec['itemsTotal']:

                record_index = record_index + 1

                # select worksheet
                wsDetails = wb['Details']
                wsReference = wb['Reference']
                wsSummary = wb['Summary']

                # details worksheet

                # Heading
                wsDetails['A1'] = "QC Dispatch Report - " + str(qaLead) + " (" + str(reportStart) + " to " + str(reportEnd) + ")"
                # Booking ID
                booking_id = str(rec['_id']['mc']) + "-" + str(rec['_id']['booking_no']) + "-" +  str(rec['_id']['type']) + "-" + str(rec['misc']['qa_type'])
                wsDetails.cell(row = row_index, column = 1, value = booking_id)
                # Country
                co = next((shipment['co'] for shipment in mcTableData if shipment['item_no'] == str(item['item_no'])), "")
                country = next((country['CountryName'] for country in countryData if country['co'] == co), "")
                wsDetails.cell(row = row_index, column = 2, value = country)
                # QA
                qa = str(rec['misc']['qa_name'])
                wsDetails.cell(row = row_index, column = 3, value = qa)
                # Inspection Date
                expInspDate = datetime.strptime(str(rec['main']['inspection_date']), "%Y-%m-%d").strftime("%m/%d/%Y")
                wsDetails.cell(row = row_index, column = 4, value = expInspDate)
                # Manufacturer
                mfName = next((party['party_name'] for party in partyTableData if party['_id'] == str(rec['main']['mf_no'])), "")
                wsDetails.cell(row = row_index, column = 5, value = mfName)
                # Inspection Type
                inspType = next((item["insp_type_long"] for item in inspecTypeData if item['insp_type'] == str(rec['_id']['type'])), "")
                wsDetails.cell(row = row_index, column = 6, value = inspType)
                # Order Qty
                orderQty = orderQty + item["order_qty"]
                wsDetails.cell(row = row_index, column = 7, value = orderQty)
                # Book ship Qty
                shipQty = shipQty + item["ship_qty"]
                wsDetails.cell(row = row_index, column = 8, value = shipQty)
                # Number of PID
                totalPID = len(set(item['item_no'] for item in rec['itemsTotal']))
                wsDetails.cell(row = row_index, column = 9, value = totalPID)

                # Reference worksheet

                # Expected Inspection Date
                wsReference.cell(row = record_index, column = 1, value = expInspDate)

                # Booking ID
                wsReference.cell(row = record_index, column = 2, value = booking_id)

                # Inspected By
                inspBy = str(rec['misc']['qa_type'])
                wsReference.cell(row = record_index, column = 3, value = inspBy)

                # Inspection Type
                wsReference.cell(row = record_index, column = 4, value = inspType)

                # Booking no
                booking_no = str(rec['_id']['booking_no'])
                wsReference.cell(row = record_index, column = 5, value = booking_no)

                # Supplier
                suName = next((party['party_name'] for party in partyTableData if party['_id'] == str(rec['main']['su_no'])), "")
                wsReference.cell(row = record_index, column = 6, value = suName)

                # Manufacturer
                wsReference.cell(row = record_index, column = 7, value = mfName)

                # PO
                po = item["po_no"]
                wsReference.cell(row = record_index, column = 8, value = po)

                # PID
                pid = item["item_no"]
                wsReference.cell(row = record_index, column = 9, value = pid)

                # Order Qty
                itemOrderQty = item["order_qty"]
                wsReference.cell(row = record_index, column = 10, value = itemOrderQty)

                # Order Qty
                itemShipQty = item["ship_qty"]
                wsReference.cell(row = record_index, column = 11, value = itemShipQty)


            
            else:

                continue


    
    wb.save(outputFile)
    wb.close()
