# -*- coding: utf-8 -*-
import uuid

from datetime import date, datetime, timedelta , timezone
from flask import jsonify, request, current_app, send_file, Blueprint
from flask import session, request, jsonify
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from io import BytesIO 
from bson.objectid import ObjectId
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.styles.alignment import Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from copy import copy
# parse date 
from dateutil import parser


import pandas as pd
import gridfs
import calendar
import smtplib
import socket
import json
import os
import re
from dotenv import load_dotenv
load_dotenv()

import checkLogged
from my_app import db,  client



booking = Blueprint('booking', __name__)

#########################################################################################################
## booking
#########################################################################################################

#db
reportMap = db["fileDirectory"]
inspectionBooking = db["inspectionBooking"]
partyTable = db["partyTable"]
metaTable = db["metaTable"]
userProfileBooking = db["userProfileBooking"]


@booking.route('/api/checkDuplicateID', methods=['POST'])
@checkLogged.check_logged
def check_duplicate_inpsection_id():
    content = request.get_json() #python data 
    _id = content['_id']
    query =  { "_id": _id}
    exists = inspectionBooking.find_one(query)
    if (exists):
        return "t",200
    else:        
        return "f",201

@booking.route('/api/save', methods=['POST'])
@checkLogged.check_logged
def save_inspection():
    content = request.get_json() #python data 

    #content = request.data # json data 
    _id = content['_id']
    
    items = content['items']
    itemsTotal = content['itemsTotal'] 
    poList = content['poList'] 
    main = content['main']
    misc = content['misc']
    update_history = content['updateHistory']    

    local_time = datetime.strptime(content['localTime'], "%a %b %d %Y %H:%M:%S")
    
    # get the current time and convert to string
    updated_time = datetime.now(timezone.utc).ctime()
    
    if (session.get("email")):
         updated_by = session.get("email")
    else:
         updated_by = "vincent.cheng@macys.com"   
    
    # Convert dictionary object into string, this is for change tracking only
    misc_str = json.dumps(misc)   
    
    #main['inspection_date'] = parser.parse(main['inspection_date'])
    main['inspection_date'] = main['inspection_date']

    ## Check only allow 14 days booking
    allow_date_start = datetime.strptime(local_time.strftime("%Y-%m-%d"), "%Y-%m-%d") + timedelta(days=(6 - datetime.strptime(local_time.strftime("%Y-%m-%d"), "%Y-%m-%d").weekday()) % 7)
    allow_date_end = datetime.strptime(local_time.strftime("%Y-%m-%d"), "%Y-%m-%d") + timedelta(days=(6 - datetime.strptime(local_time.strftime("%Y-%m-%d"), "%Y-%m-%d").weekday()) % 7) + timedelta(days=14)
    exp_insp_date = datetime.strptime(main['inspection_date'], "%Y-%m-%d")


    if (exp_insp_date >= allow_date_end) or (exp_insp_date < allow_date_start) :
        return "Inspection Date (B) error, expected inspection date should beyond 2 weeks", 202
        
    if (update_history ==[]) :        
        update_current = { "id" :  str(uuid.uuid4()), "misc" :misc_str,  "updated_by" : updated_by, "updated_time" : updated_time, "updated_mode" : "create"}
    else:
        update_current = { "id" :  str(uuid.uuid4()), "misc" :misc_str,  "updated_by" : updated_by, "updated_time" : updated_time, "updated_mode" : "update"}

    update_history.append(update_current)    
       
    new_content = { "_id" : _id, "main" : main,  "misc" : misc,   "items" : items, "itemsTotal" : itemsTotal, 
    "poList" : poList,   "update_history" : update_history }
  
    col = db["inspectionBooking"]

    query =  { "_id": _id}
    exists = col.find_one(query)


    ## convert datetime string of change tracking into datetime object in Mongodb 
    for hist in update_history:
        hist['updated_time'] = parser.parse( hist['updated_time'])

    
    if (exists):
        change =  { "$set":  {  "main" : main, "misc": misc, "items" : items, "itemsTotal" : itemsTotal, "poList" : poList,  "update_history": update_history} }    # change     
        col.update_one(query, change)
        return "ok",200
    else:        
        x = col.insert_one(new_content)
        #print(x.inserted_id)
        return "ok",200
        

def isLegitAPI(mc_no): 
    isFound = False        
    mcTable = session.get("mcTable")
    for rec in mcTable:
        if rec['mc_no'] == mc_no:
            #print ("su, mf = {0}, {1}".format(rec["su_no"], rec["mf_no"]))
            isFound = True     
            break 
            
    return (isFound)

@booking.route('/api/search',methods=['POST'])
@checkLogged.check_logged
def search_inspection():
        
    content = request.get_json() #python data     
    ##print("Search content",content)
    _id = content['_id']
    
    col = db["inspectionBooking"]

    query =  { "_id": _id}
    results = col.find_one(query)        

    if (results):
       return  jsonify(results), 200 
    else:
       return "Error", 404 

@booking.route('/api/delete',methods=['POST'])
@checkLogged.check_logged
def delete_inspection():
        
    content = request.get_json() #python data     
    ##print("Search content",content)
     
    try:  
        
        col = db["inspectionBooking"]
        delete_log_col = db["Delete_Log"]
        _id = content['_id']

        query =  { "_id": _id}

        ## add to delete_log in Mongo
        updated_by = "development@heroku" if session.get("email") == None else session.get("email")               
        existing_inspectionBooking = col.find_one(query)
        delete_log_col.insert_one( { "_id": str(uuid.uuid4()), 
        "updated_by" : updated_by, "time":datetime.now(timezone.utc),"doc_type": "inspectionBooking", "rec" : existing_inspectionBooking })

        result = col.delete_one(query)
        if result.deleted_count > 0:
            return  "OK", 200 
        else:
            return  "Not OK", 400 
    
    #except mongoengine.errors.OperationError:           
    except Exception as e: 
        print("error", e)
        return e, 400 

@booking.route('/api/searchInspByMC',methods=['POST'])
@checkLogged.check_logged
def searchInspMC():
        
    content = request.get_json() #python data     
    mc = content['mc'] 
 
    col = db["inspectionBooking"]
    
    ## Reserverd for later use
    # search = []
    # for _filter in session['mfList']:
    #     search.append(   {  '$and': [ { 'main.su_no': { '$eq': _filter['SU'] } }, { 'main.mf_no' : { '$eq': _filter['MF'] } } ]  } )        

    query =   {
       "_id.mc" : {
       "$regex": mc,
       "$options" :'i' # case-insensitive
       } }
     
    #print(search, '$$search')     

    ##results = col.find(query).limit(5)    
    ## returns 10 at a time
    ##  to access print(rec["_id"]["mc"])    
    results = col.find(query).limit(60)


    id_array = []
    for result in results:
        ##'id' : {uuid.uuid4()
        rec = { '_id' : result['_id'], 'main' : result['main'] }
        id_array.append(rec)        
        
    if (results):
       return  jsonify(id_array), 200 
    else:
       return "Error", 404 


 

@booking.route('/api/getMCtable',methods=['POST'])
@checkLogged.check_logged
def getMCtable():           

  
    content = request.get_json() #python data     
    su = content['su']
    mf  = content['mf']        
     
    #get MCs for the SU and MF only
    query =  {  '$and': [ { 'su_no': { '$eq': su } }, { 'mf_no' : { '$eq': mf } } ]  }          
    col = db["mcTable"]     
                    
    mc_array = []     
    results = col.find(query)
    for result in results:
        result["_id"] = str(result["_id"])
        mc_array.append(result) 
    
    session["mcTable"] = mc_array    

    if len(mc_array) > 0: 
        return  jsonify(mc_array), 200    


####################################################################################
#  Genearte Excel Report - Start 
####################################################################################

def bookingRecordFormatting(AllbookingRecord, localTime):

    ## Get PartyTable for su name and mf name display:
    partyTableData = list(partyTable.find({}))
    ## Get Inspection Type
    inspecTypeData = list(metaTable.find_one({'category': "inspType"})["selectionList"])
    ## Get report start and end date
    reportStart = (localTime + timedelta(days=(7 - localTime.weekday() - 1))).strftime('%m/%d')
    reportEnd = ((localTime + timedelta(days=(6 - localTime.weekday() + 6))) + timedelta(days=7)).strftime('%m/%d')

    bookingDtl = [ ]

    for lst in AllbookingRecord:

        for n in range(len(lst["itemsTotal"])):
            detail = {
                    "bookingID": str(lst["_id"].get("mc")) + "-" + str(lst["_id"].get("booking_no")) + "-" + str(lst["_id"].get("type")) + "-" + str(lst["misc"].get("qa_type")),
                    "expInspDate": datetime.strptime(lst["main"].get("inspection_date"), "%Y-%m-%d").strftime("%m/%d/%Y"),
                    "inspBy": str(lst["misc"].get("qa_type")),
                    "inspType": next((item["insp_type_long"] for item in inspecTypeData if item['insp_type'] == str(lst["_id"].get("type"))), ""),
                    "bookingNo": str(lst["_id"].get("booking_no")),
                    "suNo": str(lst["main"].get("su_no")),
                    "su": next((party['party_name'] for party in partyTableData if party['_id'] == str(lst["main"].get("su_no"))), ""),
                    "mfNo": str(lst["main"].get("mf_no")),
                    "mf": next((party['party_name'] for party in partyTableData if party['_id'] == str(lst["main"].get("mf_no"))), ""),
                    "mclist": ", ".join(set(item['mc_no'] for item in lst["itemsTotal"])),
                    "noOfPID": len(set(item['item_no'] for item in lst["itemsTotal"])),
                    "mc": lst["itemsTotal"][n]["mc_no"],
                    "po": lst["itemsTotal"][n]["po_no"],
                    "pid": lst["itemsTotal"][n]["item_no"],
                    "orderQty": lst["itemsTotal"][n]["order_qty"],
                    "shipQty": lst["itemsTotal"][n]["ship_qty"],
                    "lastUpdate": lst['update_history'][-1]['updated_by'],
                    "totalOrderQty": sum(item['order_qty'] for item in lst['itemsTotal']),
                    "totalShipQty": sum(item['ship_qty'] for item in lst['itemsTotal']),
                    "headerTitle": "QC Booking Summary for the period " +str(reportStart) + " - " + str(reportEnd)
                    }
            bookingDtl.append(detail)

    return bookingDtl


def genReport(ws, cells, bookingKey):

    if ws.title == "Details":

        # Detail Version
        for n in range(len(bookingKey)):
            for excel_range in cells:
                # Static information not included the office display and leave entry
                col_index = column_index_from_string(coordinate_from_string(cells[excel_range])[0])
                row_index = (coordinate_from_string(cells[excel_range])[1]) + n
                ws.cell(row=row_index, column=col_index, value=bookingKey[n][excel_range])
                ws.cell(row=row_index, column=col_index).border = Border(left=borders.Side(border_style='thin', color="FF000000", style=None), 
                                                                        right=borders.Side(border_style='thin', color="FF000000", style=None), 
                                                                        top=borders.Side(border_style='thin', color="FF000000", style=None),
                                                                        bottom=borders.Side(border_style='thin', color="FF000000", style=None))
                ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal='center', vertical='center')
    
    if ws.title == "Summary":

        # sort the list by bookingID
        unique_items = []
        unique_index = []
        summary_row = -1

        for index, item in enumerate(bookingKey, start=0):
            if item["bookingID"] not in unique_items:
                unique_items.append(item["bookingID"])
                unique_index.append(index)

        # Summary Version
        for n in range(len(bookingKey)):
            if n in unique_index:
                summary_row = summary_row + 1
                for excel_range in cells:
                    if excel_range != "headerTitle":
                        # Static information not included the office display and leave entry
                        col_index = column_index_from_string(coordinate_from_string(cells[excel_range])[0])
                        row_index = (coordinate_from_string(cells[excel_range])[1]) + summary_row
                        ws.cell(row=row_index, column=col_index, value=bookingKey[n][excel_range])
                        ws.cell(row=row_index, column=col_index).border = Border(left=borders.Side(border_style='thin', color="FF000000", style=None), 
                                                                                right=borders.Side(border_style='thin', color="FF000000", style=None), 
                                                                                top=borders.Side(border_style='thin', color="FF000000", style=None),
                                                                                bottom=borders.Side(border_style='thin', color="FF000000", style=None))
                        ws.cell(row=row_index, column=col_index).alignment = Alignment(horizontal='center', vertical='center')
                    # Handle Summary Heading
                    elif excel_range == "headerTitle":
                        col_index = column_index_from_string(coordinate_from_string(cells["headerTitle"])[0])
                        row_index = (coordinate_from_string(cells["headerTitle"])[1])
                        ws.cell(row=row_index, column=col_index, value=bookingKey[n][excel_range])
            else:
                continue





#Version 01  1/9/23
@booking.route('/api/printreport',methods=['POST'])
@checkLogged.check_logged
def printreport():

    # Get Report Type
    try:
        report_type = json.loads(request.headers['reportType'])
    except:
        return jsonify({"error_message" : "Sorry, we failed to generate Booking Report"}), 501
    
    if report_type == "Single":
        try:
            # Single report for specifc booking id only
            inspectionID = json.loads(request.headers['inspectionID'])
            inspectionID = {'inspectionID': [inspectionID['inspectionID']]}
            localTime = json.loads(request.headers['localTime'])
        except:
            return jsonify({"error_message" : "Sorry, we failed to generate Booking Report"}), 501
        
    elif report_type == "Multiple":
        # Get email from user profile to get SU MF list to print booking report in next 7-14 days 
        userProfile = json.loads(request.headers['userProfile'])
        email = userProfile["email"]
        mf_list = userProfileBooking.find_one({"email" : { '$eq' : email}})["mf_list"]

        # Get local Time for report period
        localTime = json.loads(request.headers['localTime'])

        # Get Booking ID BY SU MF
        inspectionID = {'inspectionID': []}
        for n in range (len(mf_list)):
            result = inspectionBooking.find( {"main.su_no":{ '$eq': str(mf_list[n]['SU'])}, 
                                              "main.mf_no":{ '$eq': str(mf_list[n]['MF'])}
                                              } )
            for row in result:
                record = {            
                    'mc': row['_id']['mc'],
                    'booking_no': row['_id']['booking_no'],
                    'type': row['_id']['type']
                    }
                inspectionID['inspectionID'].append(record)        

    # Make Report Period
    localTime = datetime.strptime(localTime, "%a %b %d %Y %H:%M:%S").date()
    reportStart = (localTime + timedelta(days=(7 - localTime.weekday() - 1)))
    reportEnd = (localTime + timedelta(days=(6 - localTime.weekday() + 6)))
        
    
    # Build the data structure for output
    # Get Record from MongoDB by using the request header from frontend
    AllbookingRecord = [ ]
    for x in range ( len (inspectionID['inspectionID']) ):
        mc = inspectionID['inspectionID'][x]['mc']
        booking_no = inspectionID['inspectionID'][x]['booking_no']
        insptype = inspectionID['inspectionID'][x]['type']

        if report_type == "Single":

            bookingRecord = inspectionBooking.find_one ( {"_id.mc" : { '$eq' : mc} , "_id.booking_no": { '$eq' :booking_no} , "_id.type": { '$eq' : insptype}} )

        elif report_type == "Multiple":

            bookingRecord = inspectionBooking.find_one ( {"_id.mc" : { '$eq' : mc} , "_id.booking_no": { '$eq' :booking_no} , "_id.type": { '$eq' : insptype}, "main.inspection_date": {"$gte": reportStart.strftime("%Y-%m-%d"),"$lte": reportEnd.strftime("%Y-%m-%d") } } )

        if bookingRecord is not None:
            AllbookingRecord.append(bookingRecord)

    bookingDetail = bookingRecordFormatting(AllbookingRecord, localTime)

    # Find Booking Report in MongoDB
    rpt = reportMap.find({"file.fileName": {"$regex": "BookingReport"}})

    # Get all worksheet if the workbook more than 1 worksheet required
    rptlist = []
    for result in rpt:
        rptlist.append(result) 
    fileObject = []

    # Get all worksheet file object in MongoDB
    for rec in rptlist:
        fileObject.append(rec['file']['fileObj'])

    # Check all file object, there should be only 1 file object if more than 1 worksheet required
    if (len(set(fileObject))) != 1 :
        return jsonify({"error_message" :"Something went wrong on fileDirectory Setting in MongoDB!"}), 502
    
    # Get file name in MongoDB by using gridFS form
    fs = gridfs.GridFS(db)
    wb = load_workbook(filename=BytesIO(fs.get(ObjectId(fileObject[0])).read()))

    # Generate file by passing worksheet name, worksheet cell value and all the variable report need:
    # Pass to genReport function
    for rec in rptlist:
        result = genReport(wb[rec['file']['wsName']], rec["cell"], bookingDetail)
    
    # Output 
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    wb.close()            

    return send_file(out,  attachment_filename='a_file.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

####################################################################################
#  Genearte Excel Report - End
####################################################################################
 



